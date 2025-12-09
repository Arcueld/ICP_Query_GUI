import sys
import asyncio
import os
import json
import yaml
import ymicp

from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QGridLayout, QLabel, QLineEdit, 
                             QPushButton, QTextEdit, QTabWidget, QGroupBox,
                             QComboBox, QCheckBox, QSpinBox, QMessageBox,
                             QProgressBar, QTableWidget, QTableWidgetItem,
                             QHeaderView, QStatusBar,
                             QFileDialog, QDialog)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from load_config import config
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

QUERY_METHODS = {
    "APP": "ymApp",
    "网站": "ymWeb",
    "小程序": "ymMiniApp",
    "快应用": "ymKuaiApp",
    "黑名单APP": "bymApp",
    "黑名单网站": "bymWeb",
    "黑名单小程序": "bymMiniApp",
    "黑名单快应用": "bymKuaiApp",
}


def execute_query(icp_instance, query_type, target):
    """根据查询类型执行对应的异步查询方法"""
    method_name = QUERY_METHODS.get(query_type)
    if method_name and hasattr(icp_instance, method_name):
        coro = getattr(icp_instance, method_name)(target)
        return icp_instance.run_async(coro)
    return {"code": 400, "message": "不支持的查询类型"}

class QueryWorker(QThread):
    result_ready = pyqtSignal(dict)
    progress_update = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, query_type, query_text):
        super().__init__()
        self.query_type = query_type
        self.query_text = query_text
        self.icp = None
        
    def run(self):
        try:
            self.icp = ymicp.beian()
            
            result = execute_query(self.icp, self.query_type, self.query_text)
            
            self.result_ready.emit(result)
            
        except Exception as e:
            self.error_occurred.emit(f"查询失败: {str(e)}")
        finally:
            if self.icp:
                self.icp.cleanup()

class BatchQueryWorker(QThread):
    result_ready = pyqtSignal(dict)
    progress_update = pyqtSignal(str, int, int)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, query_type, query_targets):
        super().__init__()
        self.query_type = query_type
        self.query_targets = query_targets
        self.icp = None
        self.batch_results = []
        system_cfg = getattr(config, 'system', object())
        max_workers = getattr(system_cfg, 'batch_thread_workers', 4)
        self.max_workers = max(1, min(16, int(max_workers))) if isinstance(max_workers, (int, float)) else 4
        
    def run(self):
        total_targets = len(self.query_targets)
        if total_targets == 0:
            self.batch_results = []
            self.result_ready.emit(self._generate_summary_result(0, 0))
            return
        
        successful_queries = 0
        failed_queries = 0
        completed = 0
        results_buffer = [None] * total_targets
        
        try:
            worker_count = min(self.max_workers, total_targets)
            with ThreadPoolExecutor(max_workers=worker_count) as executor:
                future_map = {
                    executor.submit(self._execute_single_target, idx, target): idx
                    for idx, target in enumerate(self.query_targets)
                }
                
                for future in as_completed(future_map):
                    idx = future_map[future]
                    target = self.query_targets[idx]
                    try:
                        query_result = future.result()
                        results_buffer[idx] = query_result
                        if query_result["success"]:
                            successful_queries += 1
                        else:
                            failed_queries += 1
                        status_msg = "完成"
                    except Exception as e:
                        query_result = {
                            "target": target,
                            "result": {"code": 500, "message": f"查询失败: {str(e)}"},
                            "success": False,
                            "index": idx
                        }
                        results_buffer[idx] = query_result
                        failed_queries += 1
                        status_msg = f"异常: {str(e)}"
                    
                    completed += 1
                    self.progress_update.emit(
                        f"{status_msg} {target} ({completed}/{total_targets})",
                        completed,
                        total_targets
                    )
            
            self.batch_results = [result for result in results_buffer if result]
            summary_result = self._generate_summary_result(successful_queries, failed_queries)
            self.result_ready.emit(summary_result)
        
        except Exception as e:
            self.error_occurred.emit(f"批量查询失败: {str(e)}")
    
    def _generate_summary_result(self, successful_queries, failed_queries):
        all_data = []
        for query_result in self.batch_results:
            if query_result["success"] and query_result["result"].get("params", {}).get("list"):
                for item in query_result["result"]["params"]["list"]:
                    item["_query_target"] = query_result["target"]
                    item["_query_index"] = query_result["index"]
                    all_data.append(item)
        
        summary = {
            "code": 200,
            "msg": "批量查询完成",
            "success": True,
            "params": {
                "total_queries": len(self.query_targets),
                "successful_queries": successful_queries,
                "failed_queries": failed_queries,
                "total_results": len(all_data),
                "list": all_data if all_data else [],
                "batch_results": self.batch_results
            }
        }
        return summary

    def _execute_single_target(self, index, target):
        """在独立线程中执行单个查询"""
        icp_instance = ymicp.beian()
        try:
            result = execute_query(icp_instance, self.query_type, target)
            success = result.get("code") == 200
            return {
                "target": target,
                "result": result,
                "success": success,
                "index": index
            }
        finally:
            icp_instance.cleanup()

class ConfigDialog(QDialog):
    config_saved = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("配置管理")
        self.setFixedSize(600, 500)
        self.init_ui()
        self.load_config()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        proxy_group = QGroupBox("代理配置")
        proxy_layout = QGridLayout()
        
        proxy_layout.addWidget(QLabel("启用静态代理:"), 0, 0)
        self.static_proxy_enable = QCheckBox()
        proxy_layout.addWidget(self.static_proxy_enable, 0, 1)
        
        proxy_layout.addWidget(QLabel("代理类型:"), 1, 0)
        self.proxy_type = QComboBox()
        self.proxy_type.addItems(["http", "https"])
        proxy_layout.addWidget(self.proxy_type, 1, 1)
        
        proxy_layout.addWidget(QLabel("代理地址:"), 2, 0)
        self.proxy_host = QLineEdit()
        proxy_layout.addWidget(self.proxy_host, 2, 1)
        
        proxy_layout.addWidget(QLabel("代理端口:"), 3, 0)
        self.proxy_port = QSpinBox()
        self.proxy_port.setRange(1, 65535)
        proxy_layout.addWidget(self.proxy_port, 3, 1)
        
        proxy_layout.addWidget(QLabel("用户名:"), 4, 0)
        self.proxy_username = QLineEdit()
        proxy_layout.addWidget(self.proxy_username, 4, 1)
        
        proxy_layout.addWidget(QLabel("密码:"), 5, 0)
        self.proxy_password = QLineEdit()
        self.proxy_password.setEchoMode(QLineEdit.Password)
        proxy_layout.addWidget(self.proxy_password, 5, 1)
        
        proxy_group.setLayout(proxy_layout)
        layout.addWidget(proxy_group)
        
        captcha_group = QGroupBox("验证码配置")
        captcha_layout = QGridLayout()
        
        captcha_layout.addWidget(QLabel("重试次数:"), 0, 0)
        self.captcha_retry = QSpinBox()
        self.captcha_retry.setRange(1, 10)
        captcha_layout.addWidget(self.captcha_retry, 0, 1)
        
        captcha_layout.addWidget(QLabel("重试延迟(秒):"), 1, 0)
        self.captcha_delay = QSpinBox()
        self.captcha_delay.setRange(1, 10)
        captcha_layout.addWidget(self.captcha_delay, 1, 1)
        
        captcha_group.setLayout(captcha_layout)
        layout.addWidget(captcha_group)
        
        query_group = QGroupBox("查询配置")
        query_layout = QGridLayout()
        
        query_layout.addWidget(QLabel("默认页面大小:"), 0, 0)
        self.page_size = QSpinBox()
        self.page_size.setRange(1, 100)
        self.page_size.setValue(20)
        query_layout.addWidget(self.page_size, 0, 1)
        
        query_layout.addWidget(QLabel("(每页返回的记录数，建议10-50)"), 0, 2)
        
        query_group.setLayout(query_layout)
        layout.addWidget(query_group)
        
        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("保存配置")
        self.cancel_btn = QPushButton("取消")
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        self.save_btn.clicked.connect(self.save_config)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.proxy_host.textChanged.connect(self.validate_proxy_config)
        self.proxy_port.valueChanged.connect(self.validate_proxy_config)
        
    def load_config(self):
        try:
            static_proxy = getattr(config.proxy, 'static_proxy', object())
            self.static_proxy_enable.setChecked(getattr(static_proxy, 'enable', False))
            self.proxy_type.setCurrentText(getattr(static_proxy, 'type', 'http'))
            self.proxy_host.setText(getattr(static_proxy, 'host', ''))
            self.proxy_port.setValue(getattr(static_proxy, 'port', 8080))
            self.proxy_username.setText(getattr(static_proxy, 'username', ''))
            self.proxy_password.setText(getattr(static_proxy, 'password', ''))
            
            captcha = getattr(config, 'captcha', object())
            self.captcha_retry.setValue(getattr(captcha, 'captcha_retry_times', 5))
            self.captcha_delay.setValue(getattr(captcha, 'proxy_retry_delay', 1))
            
            query = getattr(config, 'query', object())
            self.page_size.setValue(getattr(query, 'default_page_size', 20))
            
        except Exception as e:
            QMessageBox.warning(self, "警告", f"加载配置失败: {str(e)}")
    
    def save_config(self):
        try:
            config_file = "config.yml"
            if not os.path.exists(config_file):
                QMessageBox.critical(self, "错误", f"配置文件 {config_file} 不存在")
                return
                
            with open(config_file, 'r', encoding='utf-8') as f:
                config_data = yaml.safe_load(f)
            
            if 'proxy' not in config_data:
                config_data['proxy'] = {}
            if 'static_proxy' not in config_data['proxy']:
                config_data['proxy']['static_proxy'] = {}
                
            config_data['proxy']['static_proxy']['enable'] = self.static_proxy_enable.isChecked()
            config_data['proxy']['static_proxy']['type'] = self.proxy_type.currentText()
            config_data['proxy']['static_proxy']['host'] = self.proxy_host.text()
            config_data['proxy']['static_proxy']['port'] = self.proxy_port.value()
            config_data['proxy']['static_proxy']['username'] = self.proxy_username.text()
            config_data['proxy']['static_proxy']['password'] = self.proxy_password.text()
            
            # 更新验证码配置
            if 'captcha' not in config_data:
                config_data['captcha'] = {}
            config_data['captcha']['captcha_retry_times'] = self.captcha_retry.value()
            config_data['captcha']['proxy_retry_delay'] = self.captcha_delay.value()
            
            # 更新查询配置
            if 'query' not in config_data:
                config_data['query'] = {}
            config_data['query']['default_page_size'] = self.page_size.value()
            
            # 保存配置文件
            with open(config_file, 'w', encoding='utf-8') as f:
                yaml.dump(config_data, f, default_flow_style=False, allow_unicode=True, indent=2)
            
            # 发送配置保存信号
            self.config_saved.emit()
            
            QMessageBox.information(self, "成功", "配置已保存并立即生效！")
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存配置失败: {str(e)}")
    
    def validate_proxy_config(self):
        """验证代理配置"""
        if self.static_proxy_enable.isChecked():
            host = self.proxy_host.text().strip()
            port = self.proxy_port.value()
            
            if not host:
                self.save_btn.setEnabled(False)
                return
                
            if port <= 0 or port > 65535:
                self.save_btn.setEnabled(False)
                return
                
        self.save_btn.setEnabled(True)

class AboutDialog(QDialog):
    """自定义关于对话框，支持超链接"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("关于")
        self.setFixedSize(400, 200)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("ICP_Query_GUI v1.0")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        original_label = QLabel('<a href="https://github.com/HG-ha">原作者: HG-ha</a>')
        original_label.setOpenExternalLinks(True)
        original_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(original_label)
        
        fork_label = QLabel('<a href="https://github.com/Arcueld">二开: Arcueld</a>')
        fork_label.setOpenExternalLinks(True)
        fork_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(fork_label)
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
        
        self.setLayout(layout)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ICP备案查询工具 v1.0")
        self.setGeometry(100, 100, 1000, 700)
        self.init_ui()
        self.init_status_bar()
        self.init_style()
        self.update_page_size_display()
        
    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        
        query_group = QGroupBox("ICP备案查询")
        query_layout = QGridLayout()
        
        query_layout.addWidget(QLabel("查询类型:"), 0, 0)
        self.query_type = QComboBox()
        self.query_type.addItems([
            "APP", "网站", "小程序", "快应用",
            "黑名单APP", "黑名单网站", "黑名单小程序", "黑名单快应用"
        ])
        query_layout.addWidget(self.query_type, 0, 1)
        
        query_layout.addWidget(QLabel("查询内容:"), 1, 0)
        self.query_input = QTextEdit()
        self.query_input.setPlaceholderText("请输入要查询的APP名称、域名或公司名称...\n支持批量查询，每行一个查询目标")
        self.query_input.setMaximumHeight(100)
        self.query_input.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        query_layout.addWidget(self.query_input, 1, 1)
        
        self.page_size_label = QLabel("页面大小: 20")
        query_layout.addWidget(self.page_size_label, 1, 2)
        
        batch_hint = QLabel("支持批量查询：每行一个目标")
        query_layout.addWidget(batch_hint, 1, 3)
        
        self.query_btn = QPushButton("开始查询")
        self.query_btn.clicked.connect(self.start_query)
        query_layout.addWidget(self.query_btn, 1, 4)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        query_layout.addWidget(self.progress_bar, 2, 0, 1, 5)
        
        query_group.setLayout(query_layout)
        main_layout.addWidget(query_group)
        
        result_widget = QWidget()
        result_layout = QVBoxLayout()
        
        result_tabs = QTabWidget()
        
        structured_tab = QWidget()
        structured_layout = QVBoxLayout()
        
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(0)  # 初始不设置列，根据查询结果动态设置
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        structured_layout.addWidget(self.result_table)
        
        structured_tab.setLayout(structured_layout)
        result_tabs.addTab(structured_tab, "结构化结果")
        
        raw_tab = QWidget()
        raw_layout = QVBoxLayout()
        
        self.raw_result = QTextEdit()
        self.raw_result.setPlaceholderText("原始JSON结果将显示在这里...")
        raw_layout.addWidget(self.raw_result)
        
        raw_tab.setLayout(raw_layout)
        result_tabs.addTab(raw_tab, "原始JSON")
        
        result_widget.setLayout(result_layout)
        result_layout.addWidget(result_tabs)
        
        main_layout.addWidget(result_widget)
        
        toolbar_layout = QHBoxLayout()
        
        self.config_btn = QPushButton("配置管理")
        self.config_btn.clicked.connect(self.show_config)
        toolbar_layout.addWidget(self.config_btn)
        
        self.export_btn = QPushButton("导出结果")
        self.export_btn.clicked.connect(self.export_results)
        toolbar_layout.addWidget(self.export_btn)
        
        self.clear_btn = QPushButton("清空结果")
        self.clear_btn.clicked.connect(self.clear_results)
        toolbar_layout.addWidget(self.clear_btn)
        
        self.about_btn = QPushButton("关于")
        self.about_btn.clicked.connect(self.show_about)
        toolbar_layout.addWidget(self.about_btn)
        
        toolbar_layout.addStretch()
        main_layout.addLayout(toolbar_layout)
        
        central_widget.setLayout(main_layout)
        
    def init_status_bar(self):
        """初始化状态栏"""
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")
        
        self.connection_status = QLabel("连接状态: 未知")
        self.status_bar.addPermanentWidget(self.connection_status)
        
    def init_style(self):
        """初始化样式 - 使用默认样式"""
        pass
        
    def update_page_size_display(self):
        """更新页面大小显示"""
        try:
            page_size = getattr(getattr(config, 'query', object()), 'default_page_size', 20)
        except (AttributeError, TypeError):
            page_size = 20
        self.page_size_label.setText(f"页面大小: {page_size}")
        
    def start_query(self):
        """开始查询"""
        query_text = self.query_input.toPlainText().strip()
        if not query_text:
            QMessageBox.warning(self, "警告", "请输入查询内容")
            return
            
        query_type = self.query_type.currentText()
        
        query_targets = self._sanitize_query_targets(query_text)
        
        if not query_targets:
            QMessageBox.warning(self, "警告", "请输入有效的查询内容")
            return
        
        if len(query_targets) == 1:
            self._start_single_query(query_type, query_targets[0])
        else:
            self._start_batch_query(query_type, query_targets)
    
    def _start_single_query(self, query_type, query_text):
        """单个查询"""
        self.query_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        
        self.query_worker = QueryWorker(query_type, query_text)
        self.query_worker.result_ready.connect(self.on_query_result)
        self.query_worker.progress_update.connect(self.on_progress_update)
        self.query_worker.error_occurred.connect(self.on_query_error)
        self.query_worker.start()
    
    def _start_batch_query(self, query_type, query_targets):
        """批量查询"""
        self.query_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, len(query_targets))
        self.progress_bar.setValue(0)
        
        self.batch_worker = BatchQueryWorker(query_type, query_targets)
        self.batch_worker.result_ready.connect(self.on_batch_query_result)
        self.batch_worker.progress_update.connect(self.on_batch_progress_update)
        self.batch_worker.error_occurred.connect(self.on_batch_query_error)
        self.batch_worker.start()
        
    def on_query_result(self, result):
        """处理查询结果"""
        self.progress_bar.setVisible(False)
        self.query_btn.setEnabled(True)
        
        self.raw_result.setPlainText(json.dumps(result, ensure_ascii=False, indent=2))
        
        self.display_structured_result(result)
        
        self.status_bar.showMessage("查询完成")
        self.connection_status.setText("连接状态: 正常")
        
    def on_progress_update(self, message):
        """更新进度"""
        self.status_bar.showMessage(message)
        
    def on_query_error(self, error_msg):
        """处理查询错误"""
        self.progress_bar.setVisible(False)
        self.query_btn.setEnabled(True)
        
        self.status_bar.showMessage("查询失败")
        self.connection_status.setText("连接状态: 异常")
        
        QMessageBox.critical(self, "查询错误", error_msg)
    
    def on_batch_query_result(self, result):
        """处理批量查询结果"""
        self.progress_bar.setVisible(False)
        self.query_btn.setEnabled(True)
        
        self.raw_result.setPlainText(json.dumps(result, ensure_ascii=False, indent=2))
        
        self.display_structured_result(result)
        
        self.status_bar.showMessage("批量查询完成")
        self.connection_status.setText("连接状态: 正常")
    
    def on_batch_progress_update(self, message, current, total):
        """更新批量查询进度"""
        self.progress_bar.setValue(current)
        self.status_bar.showMessage(message)
    
    def on_batch_query_error(self, error_msg):
        """处理批量查询错误"""
        self.progress_bar.setVisible(False)
        self.query_btn.setEnabled(True)
        
        self.status_bar.showMessage("批量查询失败")
        self.connection_status.setText("连接状态: 异常")
        
        QMessageBox.critical(self, "批量查询错误", error_msg)
        
    def display_structured_result(self, result):
        """显示结构化结果"""
        self.result_table.setRowCount(0)
        
        if not isinstance(result, dict):
            return
        
        if result.get('code') == 404 or result.get('success') == False:
            return
        
        if 'params' in result and isinstance(result['params'], dict) and 'list' in result['params']:
            data_list = result['params']['list']
            if isinstance(data_list, list) and len(data_list) > 0:
                self.parse_data_list(data_list)
                return
            else:
                return
        
        if result.get('success', False) and result.get('code', 0) == 200:
            return
        
        return
    
    def parse_data_list(self, data_list):
        """解析数据列表 - 每行一个数据项，只显示重要信息"""
        is_batch_result = any(item.get('_query_target') for item in data_list if isinstance(item, dict))
        has_domain = any(item.get('domain') for item in data_list if isinstance(item, dict))
        
        name_header = "域名" if has_domain else "服务名称"
        
        if is_batch_result:
            self.result_table.setColumnCount(4)
            self.result_table.setHorizontalHeaderLabels(["查询目标", name_header, "单位名称", "备案号"])
        else:
            self.result_table.setColumnCount(3)
            self.result_table.setHorizontalHeaderLabels([name_header, "单位名称", "备案号"])
        
        row = 0
        for i, item in enumerate(data_list):
            if not isinstance(item, dict):
                continue
                
            self.result_table.insertRow(row)
            col = 0
            
            if is_batch_result:
                query_target = item.get('_query_target', '')
                self.result_table.setItem(row, col, QTableWidgetItem(str(query_target)))
                col += 1
            
            name_value = item.get('domain') if has_domain else item.get('serviceName', '')
            self.result_table.setItem(row, col, QTableWidgetItem(str(name_value or "")))
            col += 1
            
            unit_name = item.get('unitName', '')
            self.result_table.setItem(row, col, QTableWidgetItem(str(unit_name)))
            col += 1
            
            main_licence = item.get('mainLicence', '')
            self.result_table.setItem(row, col, QTableWidgetItem(str(main_licence)))
            
            row += 1
    
            
        
    def show_config(self):
        """显示配置对话框"""
        config_dialog = ConfigDialog(self)
        # 连接配置保存信号
        config_dialog.config_saved.connect(self.on_config_saved)
        if config_dialog.exec_() == QDialog.Accepted:
            self.status_bar.showMessage("配置已更新")
            self.update_page_size_display()
    
    def on_config_saved(self):
        """配置保存后的处理"""
        self.reload_config()
        self.update_page_size_display()
        self.status_bar.showMessage("配置已重新加载")
    
    def reload_config(self):
        """重新加载配置和查询模块"""
        global config
        import importlib
        import load_config
        importlib.reload(load_config)
        config = load_config.config
        importlib.reload(ymicp)
        
    def export_results(self):
        """导出结果"""
        if self.result_table.rowCount() == 0:
            QMessageBox.warning(self, "警告", "没有可导出的结果")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出结果", f"icp_query_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel文件 (*.xlsx);;所有文件 (*)"
        )
        
        if file_path:
            try:
                self._export_to_excel(file_path)
                QMessageBox.information(self, "成功", f"结果已导出到: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def _export_to_excel(self, file_path):
        """导出到Excel文件"""
        wb = Workbook()
        
        wb.remove(wb.active)
        
        ws_results = wb.create_sheet("查询结果")
        self._write_results_to_sheet(ws_results)
        
        wb.save(file_path)
    
    def _write_results_to_sheet(self, ws):
        """写入查询结果到工作表"""
        headers = []
        column_widths = []
        for col in range(self.result_table.columnCount()):
            header_item = self.result_table.horizontalHeaderItem(col)
            if header_item:
                headers.append(header_item.text())
            else:
                headers.append(f"列{col+1}")
            column_widths.append(len(headers[-1]))
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in range(self.result_table.rowCount()):
            for col in range(self.result_table.columnCount()):
                item = self.result_table.item(row, col)
                if item:
                    value = item.text()
                    ws.cell(row=row+2, column=col+1, value=value)
                    column_widths[col] = max(column_widths[col], len(value))
        
        for idx, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(idx)
            ws.column_dimensions[column_letter].width = min(width + 2, 50)
    
    def clear_results(self):
        """清空结果"""
        self.result_table.setRowCount(0)
        self.raw_result.clear()
        self.status_bar.showMessage("结果已清空")
        
    def show_about(self):
        """显示关于对话框"""
        about_dialog = AboutDialog(self)
        about_dialog.exec_()

    def _sanitize_query_targets(self, raw_text):
        """清洗用户输入的查询目标，去除空行、重复和无意义字符"""
        cleaned = []
        seen = set()
        max_length = 256
        for line in raw_text.splitlines():
            candidate = line.strip().strip(",;")
            if not candidate:
                continue
            if len(candidate) > max_length:
                continue
            if candidate in seen:
                continue
            cleaned.append(candidate)
            seen.add(candidate)
        return cleaned

def main():
    """主函数"""
    app = QApplication(sys.argv)
    
    # 设置应用样式
    app.setStyle('Fusion')
    
    # 创建主窗口
    window = MainWindow()
    window.show()
    
    # 运行应用
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
